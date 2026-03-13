"""Excel formatter for the Product Roadmap Agent output."""
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styles ───────────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F538D")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL = PatternFill("solid", fgColor="EEF4FB")
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER_WRAP = Alignment(horizontal="center", vertical="top", wrap_text=True)
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_PHASE_FILLS = [
    PatternFill("solid", fgColor="4472C4"),  # Phase 1 — blue
    PatternFill("solid", fgColor="ED7D31"),  # Phase 2 — orange
    PatternFill("solid", fgColor="70AD47"),  # Phase 3 — green
    PatternFill("solid", fgColor="9E480E"),  # Phase 4 — brown
]


def _header_row(ws, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER_WRAP
        cell.border = _BORDER
    ws.row_dimensions[row].height = 28


def _set_widths(ws, widths: list[int]) -> None:
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def build_roadmap_excel(data: dict) -> bytes:
    """Build a multi-sheet .xlsx from RoadmapAgent JSON output."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Overview ────────────────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov.column_dimensions["A"].width = 28
    ws_ov.column_dimensions["B"].width = 90

    title_cell = ws_ov.cell(row=1, column=1, value=data.get("roadmap_title", "Product Roadmap"))
    title_cell.fill = _HEADER_FILL
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_ov.merge_cells("A1:B1")
    ws_ov.row_dimensions[1].height = 36

    overview_fields = [
        ("Vision Statement", data.get("vision_statement", "")),
        ("Roadmap Summary", data.get("roadmap_summary", "")),
    ]
    for r_off, (label, value) in enumerate(overview_fields, 2):
        lbl = ws_ov.cell(row=r_off, column=1, value=label)
        lbl.font = Font(bold=True)
        lbl.fill = PatternFill("solid", fgColor="D6E4F0")
        lbl.border = _BORDER
        lbl.alignment = _WRAP
        val = ws_ov.cell(row=r_off, column=2, value=value)
        val.alignment = _WRAP
        val.border = _BORDER
        ws_ov.row_dimensions[r_off].height = 60

    # Release strategy block
    rs = data.get("release_strategy", {})
    if rs:
        start = len(overview_fields) + 2
        hdr = ws_ov.cell(row=start, column=1, value="Release Strategy")
        hdr.fill = _SECTION_FILL
        hdr.font = _SECTION_FONT
        hdr.border = _BORDER
        hdr.alignment = _CENTER_WRAP
        ws_ov.merge_cells(start_row=start, start_column=1, end_row=start, end_column=2)

        rs_fields = [
            ("Approach", rs.get("approach", "")),
            ("Rationale", rs.get("rationale", "")),
            ("Rollback Plan", rs.get("rollback_plan", "")),
        ]
        for i, (label, value) in enumerate(rs_fields, start + 1):
            lbl = ws_ov.cell(row=i, column=1, value=label)
            lbl.font = Font(bold=True)
            lbl.border = _BORDER
            lbl.alignment = _WRAP
            if i % 2 == 0:
                lbl.fill = _ALT_FILL
            val = ws_ov.cell(row=i, column=2, value=value)
            val.alignment = _WRAP
            val.border = _BORDER
            if i % 2 == 0:
                val.fill = _ALT_FILL

    # ── Sheet 2: Phases ──────────────────────────────────────────────────────
    ws_ph = wb.create_sheet("Phases")
    ws_ph.freeze_panes = "A2"
    _header_row(ws_ph, ["Phase", "Name", "Duration", "Objective", "Features", "Deliverables",
                         "Success Criteria", "Dependencies", "Risks"])

    for i, phase in enumerate(data.get("phases", []), 2):
        phase_fill = _PHASE_FILLS[i % len(_PHASE_FILLS)]
        alt = i % 2 == 0

        def _join(lst):
            return "\n".join(f"• {x}" for x in lst) if lst else ""

        values = [
            phase.get("phase_id", ""),
            phase.get("name", ""),
            phase.get("duration", ""),
            phase.get("objective", ""),
            ", ".join(phase.get("features", [])),
            _join(phase.get("deliverables", [])),
            _join(phase.get("success_criteria", [])),
            _join(phase.get("dependencies", [])),
            _join(phase.get("risks", [])),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_ph.cell(row=i, column=col_idx, value=val)
            cell.alignment = _WRAP
            cell.border = _BORDER
            if alt and col_idx != 1:
                cell.fill = _ALT_FILL

        phase_id_cell = ws_ph.cell(row=i, column=1)
        phase_id_cell.fill = phase_fill
        phase_id_cell.font = Font(bold=True, color="FFFFFF")
        phase_id_cell.alignment = _CENTER_WRAP

        deliverable_count = len(phase.get("deliverables", []))
        ws_ph.row_dimensions[i].height = max(40, 18 * max(deliverable_count, 2))

    _set_widths(ws_ph, [12, 28, 14, 55, 28, 55, 45, 40, 40])

    # ── Sheet 3: Milestones ──────────────────────────────────────────────────
    ws_ms = wb.create_sheet("Milestones")
    ws_ms.freeze_panes = "A2"
    _header_row(ws_ms, ["Milestone", "Target Date", "Deliverables", "KPIs Measured"])

    for i, ms in enumerate(data.get("milestones", []), 2):
        alt = i % 2 == 0
        deliverables = "\n".join(f"• {d}" for d in ms.get("deliverables", []))
        kpis = ", ".join(ms.get("kpis_measured", []))
        values = [ms.get("name", ""), ms.get("target_date_relative", ""), deliverables, kpis]
        for col_idx, val in enumerate(values, 1):
            cell = ws_ms.cell(row=i, column=col_idx, value=val)
            cell.alignment = _WRAP
            cell.border = _BORDER
            if alt:
                cell.fill = _ALT_FILL
        ws_ms.row_dimensions[i].height = max(35, 18 * max(len(ms.get("deliverables", [])), 1))

    _set_widths(ws_ms, [40, 20, 70, 30])

    # ── Sheet 4: Resources ───────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resources")
    rs_data = data.get("resource_summary", {})

    ws_res.column_dimensions["A"].width = 30
    ws_res.column_dimensions["B"].width = 60

    res_fields = [
        ("Estimated Team Size", rs_data.get("estimated_team_size", "")),
        ("Estimated Total Effort", rs_data.get("estimated_total_effort", "")),
        ("Key Roles", "\n".join(f"• {r}" for r in rs_data.get("key_roles", []))),
    ]
    hdr = ws_res.cell(row=1, column=1, value="Resource Summary")
    hdr.fill = _HEADER_FILL
    hdr.font = _HEADER_FONT
    hdr.alignment = _CENTER_WRAP
    hdr.border = _BORDER
    ws_res.merge_cells("A1:B1")
    ws_res.row_dimensions[1].height = 28

    for r_off, (label, value) in enumerate(res_fields, 2):
        lbl = ws_res.cell(row=r_off, column=1, value=label)
        lbl.font = Font(bold=True)
        lbl.fill = PatternFill("solid", fgColor="D6E4F0")
        lbl.border = _BORDER
        lbl.alignment = _WRAP
        val = ws_res.cell(row=r_off, column=2, value=value)
        val.alignment = _WRAP
        val.border = _BORDER

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
