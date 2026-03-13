"""Excel formatter for the KPI / Success Metrics Agent output."""
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styles ───────────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F538D")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL = PatternFill("solid", fgColor="EEF4FB")
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER_WRAP = Alignment(horizontal="center", vertical="top", wrap_text=True)
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_CATEGORY_FILLS = {
    "business": PatternFill("solid", fgColor="4CAF50"),
    "technical": PatternFill("solid", fgColor="2196F3"),
    "user_experience": PatternFill("solid", fgColor="9C27B0"),
    "operational": PatternFill("solid", fgColor="FF9800"),
}
_PRIORITY_FILLS = {
    "primary": PatternFill("solid", fgColor="C00000"),
    "secondary": PatternFill("solid", fgColor="70AD47"),
}


def _header_row(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER_WRAP
        cell.border = _BORDER
    ws.row_dimensions[1].height = 28


def _set_widths(ws, widths: list[int]) -> None:
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def build_kpi_excel(data: dict) -> bytes:
    """Build a multi-sheet .xlsx from KPIAgent JSON output."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: KPI Metrics ─────────────────────────────────────────────────
    ws_kpi = wb.active
    ws_kpi.title = "KPIs"
    ws_kpi.freeze_panes = "A2"

    kpi_headers = [
        "ID", "Name", "Description", "Category", "Target Value",
        "Measurement Method", "Frequency", "Baseline", "Linked Requirements", "Priority",
    ]
    _header_row(ws_kpi, kpi_headers)

    for i, kpi in enumerate(data.get("success_metrics", []), 2):
        alt = i % 2 == 0
        linked = ", ".join(kpi.get("linked_requirements", []))
        values = [
            kpi.get("id", ""),
            kpi.get("name", ""),
            kpi.get("description", ""),
            kpi.get("category", ""),
            kpi.get("target_value", ""),
            kpi.get("measurement_method", ""),
            kpi.get("frequency", ""),
            kpi.get("baseline", ""),
            linked,
            kpi.get("priority", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_kpi.cell(row=i, column=col_idx, value=val)
            cell.alignment = _WRAP
            cell.border = _BORDER
            if alt and col_idx not in (4, 10):
                cell.fill = _ALT_FILL

        cat_cell = ws_kpi.cell(row=i, column=4)
        cat_fill = _CATEGORY_FILLS.get(kpi.get("category", ""), None)
        if cat_fill:
            cat_cell.fill = cat_fill
            cat_cell.font = Font(bold=True, color="FFFFFF")
            cat_cell.alignment = _CENTER_WRAP

        pri_cell = ws_kpi.cell(row=i, column=10)
        pri_fill = _PRIORITY_FILLS.get(kpi.get("priority", ""), None)
        if pri_fill:
            pri_cell.fill = pri_fill
            pri_cell.font = Font(bold=True, color="FFFFFF")
            pri_cell.alignment = _CENTER_WRAP

        ws_kpi.row_dimensions[i].height = 45

    _set_widths(ws_kpi, [12, 30, 55, 20, 30, 55, 18, 25, 30, 14])

    # ── Sheet 2: OKRs ────────────────────────────────────────────────────────
    ws_okr = wb.create_sheet("OKRs")
    ws_okr.freeze_panes = "A2"
    _header_row(ws_okr, ["Objective", "Key Results", "Timeline"])

    for i, okr in enumerate(data.get("okrs", []), 2):
        alt = i % 2 == 0
        key_results = "\n".join(
            f"• {kr}" for kr in okr.get("key_results", [])
        )
        values = [okr.get("objective", ""), key_results, okr.get("timeline", "")]
        for col_idx, val in enumerate(values, 1):
            cell = ws_okr.cell(row=i, column=col_idx, value=val)
            cell.alignment = _WRAP
            cell.border = _BORDER
            if alt:
                cell.fill = _ALT_FILL
        kr_count = len(okr.get("key_results", []))
        ws_okr.row_dimensions[i].height = max(40, 18 * max(kr_count, 2))

    _set_widths(ws_okr, [55, 80, 25])

    # ── Sheet 3: Measurement Framework ──────────────────────────────────────
    ws_fw = wb.create_sheet("Framework")
    ws_fw.column_dimensions["A"].width = 32
    ws_fw.column_dimensions["B"].width = 80

    mf = data.get("measurement_framework", {})
    hdr = ws_fw.cell(row=1, column=1, value="Measurement Framework")
    hdr.fill = _HEADER_FILL
    hdr.font = _HEADER_FONT
    hdr.alignment = _CENTER_WRAP
    hdr.border = _BORDER
    ws_fw.merge_cells("A1:B1")
    ws_fw.row_dimensions[1].height = 28

    fw_rows = [
        ("Reporting Cadence", mf.get("reporting_cadence", "")),
        ("Dashboard Recommendations", "\n".join(f"• {d}" for d in mf.get("dashboard_recommendations", []))),
        ("Alerting Thresholds", "\n".join(f"• {a}" for a in mf.get("alerting_thresholds", []))),
        ("Review Process", mf.get("review_process", "")),
        ("Summary", data.get("success_criteria_summary", "")),
    ]
    for r_off, (label, value) in enumerate(fw_rows, 2):
        lbl = ws_fw.cell(row=r_off, column=1, value=label)
        lbl.font = Font(bold=True)
        lbl.fill = PatternFill("solid", fgColor="D6E4F0") if r_off % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        lbl.border = _BORDER
        lbl.alignment = _WRAP
        val = ws_fw.cell(row=r_off, column=2, value=value)
        val.alignment = _WRAP
        val.border = _BORDER
        ws_fw.row_dimensions[r_off].height = 50

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
