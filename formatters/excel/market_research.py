"""Excel formatter for the Market Research Agent output."""
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Shared styles ────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F538D")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL = PatternFill("solid", fgColor="EEF4FB")
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER_WRAP = Alignment(horizontal="center", vertical="top", wrap_text=True)
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_IMPACT_FILLS = {
    "high": PatternFill("solid", fgColor="FF6B6B"),
    "medium": PatternFill("solid", fgColor="FFA941"),
    "low": PatternFill("solid", fgColor="B2F5B2"),
}
_MATURITY_FILLS = {
    "emerging": PatternFill("solid", fgColor="9C27B0"),
    "growing": PatternFill("solid", fgColor="2196F3"),
    "mature": PatternFill("solid", fgColor="4CAF50"),
    "declining": PatternFill("solid", fgColor="9E9E9E"),
}


def _header_row(ws, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER_WRAP
        cell.border = _BORDER
    ws.row_dimensions[row].height = 28


def _data_row(ws, values: list, row: int, alt: bool = False) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = _WRAP
        cell.border = _BORDER
        if alt:
            cell.fill = _ALT_FILL


def _set_widths(ws, widths: list[int]) -> None:
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ─────────────────────────────────────────────────────────────────────────────

def build_market_research_excel(data: dict) -> bytes:
    """Build a multi-sheet .xlsx from MarketResearchAgent JSON output."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Overview ────────────────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = "Overview"

    ws_ov.merge_cells("A1:B1")
    title_cell = ws_ov["A1"]
    title_cell.value = "Market Research Overview"
    title_cell.fill = _HEADER_FILL
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_ov.row_dimensions[1].height = 34

    overview_rows = [
        ("Industry Context", data.get("industry_context", "")),
        ("Target Market Size", data.get("target_market_size", "")),
    ]
    for r_offset, (label, value) in enumerate(overview_rows, 2):
        lbl = ws_ov.cell(row=r_offset, column=1, value=label)
        lbl.font = Font(bold=True)
        lbl.fill = PatternFill("solid", fgColor="D6E4F0")
        lbl.border = _BORDER
        lbl.alignment = _WRAP
        val = ws_ov.cell(row=r_offset, column=2, value=value)
        val.alignment = _WRAP
        val.border = _BORDER

    # Strategic Recommendations
    recs = data.get("strategic_recommendations", [])
    if recs:
        start = len(overview_rows) + 2
        hdr = ws_ov.cell(row=start, column=1, value="Strategic Recommendations")
        hdr.font = Font(bold=True)
        hdr.fill = _SUBHEADER_FILL
        hdr.font = _SUBHEADER_FONT
        hdr.border = _BORDER
        ws_ov.merge_cells(start_row=start, start_column=1, end_row=start, end_column=2)
        for i, rec in enumerate(recs, start + 1):
            cell = ws_ov.cell(row=i, column=1, value=f"• {rec}")
            cell.alignment = _WRAP
            cell.border = _BORDER
            ws_ov.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
            if i % 2 == 0:
                cell.fill = _ALT_FILL

    ws_ov.column_dimensions["A"].width = 30
    ws_ov.column_dimensions["B"].width = 80

    # ── Sheet 2: Market Trends ───────────────────────────────────────────────
    ws_t = wb.create_sheet("Market Trends")
    ws_t.freeze_panes = "A2"
    _header_row(ws_t, ["Trend", "Relevance", "Impact"])
    for i, t in enumerate(data.get("market_trends", []), 2):
        alt = i % 2 == 0
        _data_row(ws_t, [t.get("trend", ""), t.get("relevance", ""), t.get("impact", "")], i, alt)
        impact_cell = ws_t.cell(row=i, column=3)
        impact_fill = _IMPACT_FILLS.get(t.get("impact", ""), None)
        if impact_fill:
            impact_cell.fill = impact_fill
            impact_cell.font = Font(bold=True)
            impact_cell.alignment = _CENTER_WRAP
    _set_widths(ws_t, [35, 70, 15])

    # ── Sheet 3: Competitive Landscape ──────────────────────────────────────
    ws_c = wb.create_sheet("Competitive Landscape")
    ws_c.freeze_panes = "A2"
    _header_row(ws_c, ["Solution Category", "Key Players", "Differentiation Opportunities"])
    for i, entry in enumerate(data.get("competitive_landscape", []), 2):
        players = ", ".join(entry.get("key_players", []))
        _data_row(ws_c,
                  [entry.get("category", ""), players, entry.get("differentiation_opportunities", "")],
                  i, i % 2 == 0)
    _set_widths(ws_c, [35, 45, 70])

    # ── Sheet 4: Technology Considerations ──────────────────────────────────
    ws_tech = wb.create_sheet("Technology")
    ws_tech.freeze_panes = "A2"
    _header_row(ws_tech, ["Technology", "Maturity", "Recommendation"])
    for i, tech in enumerate(data.get("technology_considerations", []), 2):
        _data_row(ws_tech,
                  [tech.get("technology", ""), tech.get("maturity", ""), tech.get("recommendation", "")],
                  i, i % 2 == 0)
        mat_cell = ws_tech.cell(row=i, column=2)
        mat_fill = _MATURITY_FILLS.get(tech.get("maturity", ""), None)
        if mat_fill:
            mat_cell.fill = mat_fill
            mat_cell.font = Font(bold=True, color="FFFFFF")
            mat_cell.alignment = _CENTER_WRAP
    _set_widths(ws_tech, [35, 18, 80])

    # ── Sheet 5: Risk Factors ────────────────────────────────────────────────
    ws_r = wb.create_sheet("Risk Factors")
    ws_r.freeze_panes = "A2"
    _header_row(ws_r, ["Risk", "Likelihood", "Mitigation"])
    for i, risk in enumerate(data.get("risk_factors", []), 2):
        _data_row(ws_r,
                  [risk.get("risk", ""), risk.get("likelihood", ""), risk.get("mitigation", "")],
                  i, i % 2 == 0)
        lh_cell = ws_r.cell(row=i, column=2)
        lh_fill = _IMPACT_FILLS.get(risk.get("likelihood", ""), None)
        if lh_fill:
            lh_cell.fill = lh_fill
            lh_cell.font = Font(bold=True)
            lh_cell.alignment = _CENTER_WRAP
    _set_widths(ws_r, [55, 18, 70])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
