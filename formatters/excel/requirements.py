"""Excel formatter for the Requirements Intelligence Agent output."""
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styles ──────────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F538D")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL = PatternFill("solid", fgColor="EEF4FB")
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER_WRAP = Alignment(horizontal="center", vertical="top", wrap_text=True)
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Priority badge colours
_PRIORITY_FILLS = {
    "must-have": PatternFill("solid", fgColor="FF6B6B"),
    "should-have": PatternFill("solid", fgColor="FFA941"),
    "could-have": PatternFill("solid", fgColor="FFE066"),
    "wont-have": PatternFill("solid", fgColor="C0C0C0"),
}

# Category badge colours
_CATEGORY_FILLS = {
    "functional": PatternFill("solid", fgColor="4CAF50"),
    "non-functional": PatternFill("solid", fgColor="2196F3"),
    "constraint": PatternFill("solid", fgColor="FF9800"),
    "compliance": PatternFill("solid", fgColor="9C27B0"),
}


def _apply_header(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER_WRAP
        cell.border = _BORDER


def build_requirements_excel(requirements: list[dict]) -> bytes:
    """Build a styled .xlsx from the requirements list returned by RequirementsAgent."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requirements"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    headers = ["ID", "Title", "Description", "Category", "Priority", "Source Section"]
    _apply_header(ws, headers)

    for row_idx, req in enumerate(requirements, 2):
        alt = row_idx % 2 == 0
        values = [
            req.get("id", ""),
            req.get("title", ""),
            req.get("description", ""),
            req.get("category", ""),
            req.get("priority", ""),
            req.get("source_section", ""),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _WRAP
            cell.border = _BORDER
            if alt and col_idx not in (4, 5):  # skip badge cols for alt fill
                cell.fill = _ALT_FILL

        # Category badge colouring
        cat_cell = ws.cell(row=row_idx, column=4)
        cat_fill = _CATEGORY_FILLS.get(req.get("category", ""), None)
        if cat_fill:
            cat_cell.fill = cat_fill
            cat_cell.font = Font(bold=True, color="FFFFFF")
            cat_cell.alignment = _CENTER_WRAP

        # Priority badge colouring
        pri_cell = ws.cell(row=row_idx, column=5)
        pri_fill = _PRIORITY_FILLS.get(req.get("priority", ""), None)
        if pri_fill:
            pri_cell.fill = pri_fill
            pri_cell.font = Font(bold=True)
            pri_cell.alignment = _CENTER_WRAP

    # Set column widths
    for col_idx, width in enumerate([12, 35, 70, 22, 20, 28], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Summary stats at bottom
    if requirements:
        ws.append([])
        ws.append(["Total Requirements", len(requirements)])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
