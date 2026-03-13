"""Excel formatter for the Feature Planning Agent output.

Sheet structure mirrors the Functional Area / Feature / Sub-Features layout
shown in the product spec screenshot, using Priority as the grouping dimension.
"""
from io import BytesIO
from itertools import groupby

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styles ───────────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F538D")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Priority colours (A column group badges)
_PRIORITY_FILLS = {
    "P0": PatternFill("solid", fgColor="C00000"),  # dark red — critical
    "P1": PatternFill("solid", fgColor="E97132"),  # orange — important
    "P2": PatternFill("solid", fgColor="FFD966"),  # yellow — nice-to-have
    "P3": PatternFill("solid", fgColor="70AD47"),  # green — future
}
_PRIORITY_FONTS = {
    "P0": Font(bold=True, color="FFFFFF", size=12),
    "P1": Font(bold=True, color="FFFFFF", size=12),
    "P2": Font(bold=True, color="000000", size=12),
    "P3": Font(bold=True, color="FFFFFF", size=12),
}
_PRIORITY_LABELS = {
    "P0": "P0\nCritical",
    "P1": "P1\nImportant",
    "P2": "P2\nNice-to-Have",
    "P3": "P3\nFuture",
}
_PRIORITY_ORDER = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}

_ALT_FILL = PatternFill("solid", fgColor="EEF4FB")


def _apply_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER_WRAP
        cell.border = _BORDER
    ws.row_dimensions[1].height = 30


def build_feature_planning_excel(features: list[dict]) -> bytes:
    """Build a styled .xlsx from the features list returned by FeaturePlanningAgent."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Feature Backlog (grouped by priority) ───────────────────────
    ws = wb.active
    ws.title = "Feature Backlog"
    ws.freeze_panes = "A2"

    headers = [
        "Priority",
        "Feature ID",
        "Feature Title",
        "Description",
        "User Story",
        "Score",
        "Acceptance Criteria",
        "Linked Requirements",
    ]
    _apply_header(ws, headers)

    sorted_features = sorted(
        features,
        key=lambda f: _PRIORITY_ORDER.get(f.get("priority", ""), 99),
    )

    current_row = 2
    for priority, group in groupby(sorted_features, key=lambda f: f.get("priority", "Unknown")):
        group_list = list(group)
        group_start = current_row

        for feat in group_list:
            ac_text = "\n".join(
                f"• {ac}" for ac in feat.get("acceptance_criteria", [])
            ) or ""
            linked = ", ".join(feat.get("linked_requirements", []))

            row_values = [
                _PRIORITY_LABELS.get(priority, priority),
                feat.get("id", ""),
                feat.get("title", ""),
                feat.get("description", ""),
                feat.get("user_story", ""),
                feat.get("priority_score", ""),
                ac_text,
                linked,
            ]
            alt = current_row % 2 == 0
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = _BORDER
                cell.alignment = _WRAP
                if alt and col_idx != 1:  # skip priority col (handled separately)
                    cell.fill = _ALT_FILL

            # Estimate row height from acceptance criteria count
            ac_count = len(feat.get("acceptance_criteria", []))
            ws.row_dimensions[current_row].height = max(40, 18 * max(ac_count, 2))

            current_row += 1

        # Merge and style the Priority column for this group
        if current_row - 1 >= group_start:
            if current_row - 1 > group_start:
                ws.merge_cells(
                    start_row=group_start,
                    start_column=1,
                    end_row=current_row - 1,
                    end_column=1,
                )
            pri_cell = ws.cell(row=group_start, column=1)
            pri_cell.fill = _PRIORITY_FILLS.get(priority, PatternFill("solid", fgColor="CCCCCC"))
            pri_cell.font = _PRIORITY_FONTS.get(priority, Font(bold=True, size=12))
            pri_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    for col_idx, width in enumerate([14, 12, 30, 55, 55, 7, 55, 28], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 2: Summary by Priority ─────────────────────────────────────────
    ws2 = wb.create_sheet("Priority Summary")
    ws2.freeze_panes = "A2"
    _apply_header(ws2, ["Priority", "Count", "Feature IDs"])

    summary: dict[str, list] = {}
    for feat in features:
        p = feat.get("priority", "Unknown")
        summary.setdefault(p, []).append(feat.get("id", ""))

    for row_idx, (p, ids) in enumerate(
        sorted(summary.items(), key=lambda x: _PRIORITY_ORDER.get(x[0], 99)), 2
    ):
        ws2.cell(row=row_idx, column=1, value=p).border = _BORDER
        ws2.cell(row=row_idx, column=2, value=len(ids)).border = _BORDER
        ws2.cell(row=row_idx, column=3, value=", ".join(ids)).border = _BORDER
        pri_fill = _PRIORITY_FILLS.get(p, PatternFill("solid", fgColor="CCCCCC"))
        ws2.cell(row=row_idx, column=1).fill = pri_fill
        ws2.cell(row=row_idx, column=1).font = _PRIORITY_FONTS.get(p, Font(bold=True))
        ws2.cell(row=row_idx, column=1).alignment = _CENTER_WRAP

    for col_idx, width in enumerate([14, 10, 60], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
